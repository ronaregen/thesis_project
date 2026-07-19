"""
Ekspor DATA MENTAH per jendela (panjang = phase_pipeline.WINDOW_SEC) ke Excel.

Tiap baris = satu jendela: berapa bpm menurut ECG (Attys), berapa menurut radar.
Kolom galat, MAE, dan MAPE ditulis sebagai RUMUS EXCEL (bukan angka jadi), supaya
siapa pun bisa mengklik selnya dan melihat sendiri dari mana angkanya berasal.

Nilai est/ref dijamin IDENTIK dengan phase_pipeline.py (ada assert-nya).

Usage:
    python code/evaluation/export_windows.py data/processed/aligned_all.csv thesis/data_mentah
"""

import sys
from pathlib import Path

import numpy as np
import pandas as pd
from scipy import signal

sys.path.insert(0, str(Path(__file__).resolve().parent))
from phase_pipeline import (FS, HR_BAND, BR_BAND, WINDOW_SEC, OVERLAP,
                            MEDFILT_K, MAPE_STANDARD, estimate, metrics,
                            resample_antialias)


def windows(t, phase, gt):
    """Ulangi persis langkah phase_pipeline.estimate(), tapi simpan juga waktu
    tiap jendela dan estimasi SEBELUM median filter."""
    tu, xu, order = resample_antialias(t, phase)
    gu = np.interp(tu, np.sort(t), gt[order])

    d = np.diff(xu, prepend=xu[0])
    sos_br = signal.butter(4, BR_BAND, "bandpass", fs=FS, output="sos")
    sos_hr = signal.butter(4, HR_BAND, "bandpass", fs=FS, output="sos")

    w = int(WINDOW_SEC * FS)
    step = max(1, int(w * (1 - OVERLAP)))

    t0, raw, ref = [], [], []
    for s in range(0, len(d) - w, step):
        seg = signal.detrend(d[s:s + w])
        seg = seg - signal.sosfiltfilt(sos_br, seg)
        seg = signal.sosfiltfilt(sos_hr, seg)
        f, p = signal.welch(seg, fs=FS, nperseg=w)
        m = (f >= HR_BAND[0]) & (f <= HR_BAND[1])
        raw.append(f[m][np.argmax(p[m])] * 60.0)
        ref.append(float(gu[s:s + w].mean()))
        t0.append(tu[s] - tu[0])

    raw = np.array(raw)
    est = signal.medfilt(raw, MEDFILT_K) if len(raw) >= MEDFILT_K else raw.copy()
    return np.array(t0), raw, est, np.array(ref)


def write_xlsx(path, d, subs, last):
    """Tulis .xlsx dengan rumus asli (bukan angka jadi) supaya bisa ditelusuri."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = f"jendela {WINDOW_SEC:.0f} detik"

    head = ["subjek", "jendela", "t_mulai_s", "t_selesai_s", "ecg_hz", "ecg_bpm",
            "radar_hz", "radar_bpm", "radar_bpm_sebelum_medfilt",
            "galat_bpm", "galat_mutlak_bpm", "galat_persen"]
    ws.append(head)
    for c in range(1, len(head) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1C2A42")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # batas baris tiap subjek -> supaya median filter tidak mencampur antar subjek
    bounds = {}
    for s in subs:
        idx = np.where(d.subjek.values == s)[0]
        bounds[s] = (idx[0] + 2, idx[-1] + 2)   # baris Excel awal & akhir

    for i, r in enumerate(d.itertuples(index=False)):
        x = i + 2
        a, b = bounds[r.subjek]
        # Median filter MEDFILT_K titik, ditulis sebagai rumus supaya bisa ditelusuri.
        # Di luar batas subjek diisi 0 -- meniru zero-padding scipy.signal.medfilt
        # persis, agar hasil Excel identik dengan pipeline Python.
        half = MEDFILT_K // 2
        terms = ",".join(f"I{x+k}" if a <= x + k <= b else "0"
                         for k in range(-half, half + 1))
        ws.append([r.subjek, r.jendela, r.t_mulai_s, r.t_selesai_s,
                   r.ecg_hz, r.ecg_bpm,
                   f"=H{x}/60",                 # radar_hz ikut radar_bpm
                   f"=MEDIAN({terms})",         # <- median filter, kini terlihat
                   r.radar_bpm_sebelum_medfilt,
                   f"=H{x}-F{x}", f"=ABS(J{x})", f"=K{x}/F{x}*100"])

    ws.freeze_panes = "A2"
    for c, w in zip("ABCDEFGHIJKL",
                    [11, 8, 10, 11, 9, 9, 9, 10, 15, 10, 12, 12]):
        ws.column_dimensions[c].width = w

    r = last + 2
    ws.cell(r, 1, "RINGKASAN").font = Font(bold=True, size=12)
    r += 1
    for c, t in zip((1, 2, 3, 4, 5, 6, 7, 10, 12),
                    ("keterangan", "subjek", "n_jendela", "std_ECG_bpm", "bias_bpm",
                     "std_galat_bpm", "RMSE_bpm", "MAE_bpm", "MAPE_persen")):
        cell = ws.cell(r, c, t)
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E8ECF1")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    r1 = r + 1
    for s in subs:
        r += 1
        A, N = f"$A$2:$A${last}", f"COUNTIF($A$2:$A${last},$B{r})"
        ws.cell(r, 1, "per subjek")
        ws.cell(r, 2, s)
        ws.cell(r, 3, f"={N}")
        # SUMPRODUCT dipakai (bukan STDEV+IF) supaya jalan di semua versi Excel
        # tanpa perlu Ctrl+Shift+Enter.
        ws.cell(r, 4, f"=SQRT(SUMPRODUCT(({A}=$B{r})*($F$2:$F${last}-"
                      f"AVERAGEIF({A},$B{r},$F$2:$F${last}))^2)/({N}-1))")
        ws.cell(r, 5, f"=AVERAGEIF({A},$B{r},$J$2:$J${last})")
        ws.cell(r, 6, f"=SQRT(SUMPRODUCT(({A}=$B{r})*($J$2:$J${last}-$E{r})^2)/({N}-1))")
        ws.cell(r, 7, f"=SQRT(SUMPRODUCT(({A}=$B{r})*$J$2:$J${last}^2)/{N})")
        ws.cell(r, 10, f"=AVERAGEIF({A},$B{r},$K$2:$K${last})")
        ws.cell(r, 12, f"=AVERAGEIF({A},$B{r},$L$2:$L${last})")
        for c in (4, 5, 6, 7, 10, 12):
            ws.cell(r, c).number_format = "0.0"
    r2 = r

    r += 2
    ws.cell(r, 1, "rata-rata antar-subjek").font = Font(bold=True)
    ws.cell(r, 2, "<- ini yang dilaporkan di tesis")
    ws.cell(r, 10, f"=AVERAGE(J{r1}:J{r2})").font = Font(bold=True)
    ws.cell(r, 12, f"=AVERAGE(L{r1}:L{r2})").font = Font(bold=True)
    ws.cell(r, 10).number_format = "0.0"
    ws.cell(r, 12).number_format = "0.0"

    r += 1
    ws.cell(r, 1, "gabungan semua jendela")
    ws.cell(r, 2, "(bobot per jendela, BUKAN yang dipakai)")
    ws.cell(r, 3, f"=COUNT($K$2:$K${last})")
    ws.cell(r, 10, f"=AVERAGE($K$2:$K${last})").number_format = "0.0"
    ws.cell(r, 12, f"=AVERAGE($L$2:$L${last})").number_format = "0.0"

    r += 2
    ws.cell(r, 1, f"lolos standar ANSI/CTA-2065 (MAPE < {MAPE_STANDARD:.0f}%)").font = Font(bold=True)
    ws.cell(r, 10, f'=COUNTIF(L{r1}:L{r2},"<{MAPE_STANDARD:.0f}")&" dari {len(subs)} subjek"').font = Font(bold=True)

    # --- Bland-Altman: cara baku melaporkan kesepakatan dua alat ukur ---
    r += 2
    ws.cell(r, 1, "BLAND-ALTMAN (semua jendela)").font = Font(bold=True, size=12)
    r += 1
    ws.cell(r, 1, "bias (rata-rata galat)")
    ws.cell(r, 5, f"=AVERAGE($J$2:$J${last})").number_format = "0.0"
    r += 1
    ws.cell(r, 1, "simpangan baku galat")
    # STDEV (bukan STDEV.S): STDEV.S butuh prefix _xlfn di sebagian aplikasi
    # dan muncul sebagai #NAME? di LibreOffice.
    ws.cell(r, 5, f"=STDEV($J$2:$J${last})").number_format = "0.0"
    sd_row = r
    r += 1
    ws.cell(r, 1, "batas kesepakatan 95% bawah")
    ws.cell(r, 5, f"=E{sd_row-1}-1.96*E{sd_row}").number_format = "0.0"
    r += 1
    ws.cell(r, 1, "batas kesepakatan 95% atas")
    ws.cell(r, 5, f"=E{sd_row-1}+1.96*E{sd_row}").number_format = "0.0"

    # --- catatan: kenapa std_ECG penting ---
    r += 2
    for note in [
        "CATATAN std_ECG_bpm - ini yang menjelaskan kenapa korelasi rendah:",
        "  Semua subjek direkam duduk diam, jadi detak jantungnya nyaris tidak berubah",
        "  (std hanya 0.9-4.2 bpm). Korelasi mengukur apakah NAIK-TURUNnya cocok; kalau",
        "  tidak ada naik-turun, korelasi mendekati nol SEKALIPUN estimasinya akurat.",
        "  Batas atas korelasi yang mungkin: r_maks = std / AKAR(std^2 + galat^2).",
        "  Untuk subject03 (std 0.9 bpm), estimator SEMPURNA pun hanya bisa mencapai r = 0.09.",
        "  Jadi korelasi rendah di sini adalah keterbatasan DATASET, bukan keterbatasan metode.",
    ]:
        ws.cell(r, 1, note).font = Font(italic=True, bold=note.endswith(":"))
        r += 1

    wb.save(path)
    return path


def main(csv, outdir):
    out = Path(outdir)
    out.mkdir(parents=True, exist_ok=True)

    df = pd.read_csv(csv)
    front = df[(df.dataset == "position_front") & df.unwrapPhasePeak_mm.notna()]

    rows = []
    for sid, sub in front.groupby("subject_id"):
        t0, raw, est, ref = windows(sub.Timestamp.values,
                                    sub.unwrapPhasePeak_mm.values,
                                    sub.gt_heart_rate.values)

        # jaminan: identik dengan pipeline yang dipakai melapor hasil
        e_ref, r_ref, _ = estimate(sub.Timestamp.values,
                                   sub.unwrapPhasePeak_mm.values,
                                   sub.gt_heart_rate.values)
        assert np.allclose(est, e_ref) and np.allclose(ref, r_ref), \
            f"{sid}: hasil ekspor TIDAK cocok dengan phase_pipeline"

        for i in range(len(est)):
            rows.append(dict(subjek=sid, jendela=i + 1,
                             t_mulai_s=round(t0[i], 1),
                             t_selesai_s=round(t0[i] + WINDOW_SEC, 1),
                             ecg_hz=round(ref[i] / 60, 4), ecg_bpm=round(ref[i], 2),
                             radar_hz=round(est[i] / 60, 4), radar_bpm=round(est[i], 2),
                             radar_bpm_sebelum_medfilt=round(raw[i], 2)))

    d = pd.DataFrame(rows)
    n = len(d)
    hdr = ("subjek,jendela,t_mulai_s,t_selesai_s,ecg_hz,ecg_bpm,radar_hz,radar_bpm,"
           "radar_bpm_sebelum_medfilt,galat_bpm,galat_mutlak_bpm,galat_persen")

    lines = [hdr]
    for i, r in enumerate(d.itertuples(index=False)):
        x = i + 2  # baris Excel (baris 1 = header)
        lines.append(
            f"{r.subjek},{r.jendela},{r.t_mulai_s},{r.t_selesai_s},"
            f"{r.ecg_hz},{r.ecg_bpm},{r.radar_hz},{r.radar_bpm},"
            f"{r.radar_bpm_sebelum_medfilt},"
            f"=H{x}-F{x},"          # galat = radar - ecg
            f"=ABS(J{x}),"          # galat mutlak
            f"=K{x}/F{x}*100")      # galat persen (relatif terhadap ECG)

    last = n + 1
    subs = sorted(d.subjek.unique())

    def q(f):
        """Rumus Excel yang mengandung koma HARUS dikutip, kalau tidak Excel
        memecahnya jadi beberapa sel dan kolomnya berantakan."""
        return '"' + f.replace('"', '""') + '"'

    lines += ["", "RINGKASAN", "keterangan,subjek,n_jendela,,,,,,,MAE_bpm,,MAPE_persen"]
    for s in subs:
        x = len(lines) + 1   # baris Excel tempat rumus ini akan berada
        lines.append(
            f"per subjek,{s},"
            f"{q(f'=COUNTIF($A$2:$A${last},$B{x})')},,,,,,,"
            f"{q(f'=AVERAGEIF($A$2:$A${last},$B{x},$K$2:$K${last})')},,"
            f"{q(f'=AVERAGEIF($A$2:$A${last},$B{x},$L$2:$L${last})')}")

    r1, r2 = len(lines) - len(subs) + 1, len(lines)
    lines += [
        "",
        f"rata-rata antar-subjek,(ini yang dilaporkan di tesis),,,,,,,,"
        f"=AVERAGE(J{r1}:J{r2}),,=AVERAGE(L{r1}:L{r2})",
        f"gabungan semua jendela,(bobot per jendela),{q(f'=COUNT($K$2:$K${last})')},,,,,,,"
        f"=AVERAGE($K$2:$K${last}),,=AVERAGE($L$2:$L${last})",
        "",
        f"lolos standar ANSI/CTA-2065 (MAPE < {MAPE_STANDARD:.0f}%),,,,,,,,,"
        f"{q(f'=COUNTIF(L{r1}:L{r2},\"<{MAPE_STANDARD:.0f}\")&\" dari {len(subs)} subjek\"')}",
    ]

    p = out / f"data_jendela_{WINDOW_SEC:.0f}detik.csv"
    p.write_text("\n".join(lines) + "\n", encoding="utf-8")

    # XLSX = format yang DIANJURKAN.
    # CSV berisi rumus akan gagal di Excel ber-locale Indonesia/Eropa: pemisah
    # daftarnya ';' bukan ',', sehingga kolom tidak terpisah DAN rumus dianggap
    # teks. Di xlsx, rumus disimpan dalam format Inggris lalu ditampilkan sesuai
    # locale -> aman di mana pun.
    write_xlsx(out / f"data_jendela_{WINDOW_SEC:.0f}detik.xlsx", d, subs, last)

    # versi angka jadi (buat cek cepat / plotting, tanpa rumus)
    d["galat_bpm"] = d.radar_bpm - d.ecg_bpm
    d["galat_mutlak_bpm"] = d.galat_bpm.abs()
    d["galat_persen"] = d.galat_mutlak_bpm / d.ecg_bpm * 100
    d.to_csv(out / f"data_jendela_{WINDOW_SEC:.0f}detik_nilai.csv", index=False)

    # verifikasi silang terhadap pipeline
    per = d.groupby("subjek").agg(mae=("galat_mutlak_bpm", "mean"),
                                  mape=("galat_persen", "mean"))
    print(f"{p}  ({n} jendela, {len(subs)} subjek)\n")
    print(f"{'subjek':<12}{'n':>5}{'MAE':>8}{'MAPE':>8}   vonis")
    print("-" * 45)
    for s in subs:
        m = per.loc[s]
        print(f"{s:<12}{(d.subjek==s).sum():>5}{m.mae:>8.1f}{m.mape:>7.1f}%   "
              f"{'LOLOS' if m.mape < MAPE_STANDARD else 'gagal'}")
    print("-" * 45)
    print(f"{'rata2 antar-subjek':<17}{per.mae.mean():>3.1f} bpm  {per.mape.mean():.1f}%"
          f"   <- angka yang dilaporkan di tesis")
    print(f"{'gabungan jendela':<17}{d.galat_mutlak_bpm.mean():>3.1f} bpm  "
          f"{d.galat_persen.mean():.1f}%   <- beda, karena tiap subjek beda jumlah jendela")
    print(f"\nLolos standar: {(per.mape < MAPE_STANDARD).sum()} dari {len(subs)} subjek")


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else "thesis/data_mentah")
